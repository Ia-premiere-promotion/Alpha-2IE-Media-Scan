"""
Route pour l'export complet de la base de données en Excel
"""
from flask import Blueprint, send_file, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from io import BytesIO
import sys
from pathlib import Path

# Ajouter le path parent pour les imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from supabase_client import get_supabase_client

export_bp = Blueprint('export', __name__, url_prefix='/api/export')


@export_bp.route('/database', methods=['GET'])
@jwt_required()
def export_database():
    """
    Exporte toute la base de données en Excel
    Chaque table = 1 feuille dans le fichier Excel
    """
    try:
        # Vérifier les permissions (admin uniquement)
        current_user = get_jwt_identity()
        supabase = get_supabase_client()
        
        # Créer un nouveau classeur Excel
        wb = Workbook()
        wb.remove(wb.active)  # Supprimer la feuille par défaut
        
        # Style pour les en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Liste des tables à exporter
        tables_config = [
            {
                'name': 'Articles',
                'table': 'articles',
                'columns': ['id', 'media', 'titre', 'contenu', 'url', 'date_publication', 
                           'categorie', 'sentiment', 'type_source', 'plateforme', 'created_at']
            },
            {
                'name': 'Médias',
                'table': 'medias',
                'columns': ['id', 'nom', 'type', 'url', 'description', 'categorie', 
                           'actif', 'page_facebook', 'created_at']
            },
            {
                'name': 'Utilisateurs',
                'table': 'users',
                'columns': ['id', 'nom', 'prenom', 'email', 'role', 'actif', 'created_at']
            },
            {
                'name': 'Alertes',
                'table': 'alerts',
                'columns': ['id', 'media', 'type_alerte', 'severite', 'message', 
                           'valeur_actuelle', 'seuil', 'is_resolved', 'created_at']
            },
            {
                'name': 'Scores Déontologie',
                'table': 'deontology_scores',
                'columns': ['id', 'article_id', 'score_global', 'objectivite', 'exactitude',
                           'equilibre', 'transparence', 'respect_personne', 'respect_vie_privee',
                           'independance', 'responsabilite', 'created_at']
            }
        ]
        
        # Exporter chaque table
        for table_config in tables_config:
            try:
                # Récupérer les données
                response = supabase.table(table_config['table']).select('*').execute()
                data = response.data
                
                if not data:
                    continue  # Passer si la table est vide
                
                # Créer une nouvelle feuille
                ws = wb.create_sheet(title=table_config['name'])
                
                # Écrire les en-têtes
                columns = table_config['columns']
                for col_idx, column in enumerate(columns, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=column.upper())
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Écrire les données
                for row_idx, record in enumerate(data, start=2):
                    for col_idx, column in enumerate(columns, start=1):
                        value = record.get(column, '')
                        
                        # Formater les valeurs
                        if value is None:
                            value = ''
                        elif isinstance(value, (dict, list)):
                            value = str(value)
                        
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Ajuster la largeur des colonnes
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Max 50 caractères
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                print(f"✅ Table '{table_config['name']}' exportée : {len(data)} lignes")
                
            except Exception as e:
                print(f"⚠️ Erreur export table '{table_config['name']}': {e}")
                continue
        
        # Créer un buffer en mémoire
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nom du fichier avec timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'MEDIA_SCAN_Database_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Erreur export database: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@export_bp.route('/articles', methods=['GET'])
@jwt_required()
def export_articles_only():
    """
    Exporte uniquement les articles en Excel
    """
    try:
        supabase = get_supabase_client()
        
        # Récupérer tous les articles
        response = supabase.table('articles').select('*').order('date_publication', desc=True).execute()
        data = response.data
        
        # Créer le classeur
        wb = Workbook()
        ws = wb.active
        ws.title = 'Articles'
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        
        # En-têtes
        headers = ['ID', 'Média', 'Titre', 'Contenu', 'URL', 'Date Publication', 
                  'Catégorie', 'Sentiment', 'Type Source', 'Plateforme', 'Créé le']
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Données
        for row_idx, article in enumerate(data, start=2):
            ws.cell(row=row_idx, column=1, value=article.get('id', ''))
            ws.cell(row=row_idx, column=2, value=article.get('media', ''))
            ws.cell(row=row_idx, column=3, value=article.get('titre', ''))
            ws.cell(row=row_idx, column=4, value=article.get('contenu', ''))
            ws.cell(row=row_idx, column=5, value=article.get('url', ''))
            ws.cell(row=row_idx, column=6, value=article.get('date_publication', ''))
            ws.cell(row=row_idx, column=7, value=article.get('categorie', ''))
            ws.cell(row=row_idx, column=8, value=article.get('sentiment', ''))
            ws.cell(row=row_idx, column=9, value=article.get('type_source', ''))
            ws.cell(row=row_idx, column=10, value=article.get('plateforme', ''))
            ws.cell(row=row_idx, column=11, value=article.get('created_at', ''))
        
        # Ajuster largeurs
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Sauvegarder
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'MEDIA_SCAN_Articles_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Erreur export articles: {e}")
        return jsonify({'error': str(e)}), 500
